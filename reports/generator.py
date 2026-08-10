"""
Report generator - HTML, JSON, Excel, PDF, and SARIF v2.1.0 (CI/CD integration)
"""
import io
import json
import html
import re
from datetime import datetime
from typing import Optional, Dict, Any, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable, KeepTogether
    )
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.pdfgen import canvas
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


SEVERITY_COLORS = {
    "critical": "#dc2626",
    "high": "#ea580c",
    "medium": "#d97706",
    "low": "#2563eb",
    "info": "#64748b",
}

SEVERITY_HEX = {
    "critical": colors.HexColor("#dc2626") if HAS_REPORTLAB else "#dc2626",
    "high": colors.HexColor("#ea580c") if HAS_REPORTLAB else "#ea580c",
    "medium": colors.HexColor("#d97706") if HAS_REPORTLAB else "#d97706",
    "low": colors.HexColor("#2563eb") if HAS_REPORTLAB else "#2563eb",
    "info": colors.HexColor("#64748b") if HAS_REPORTLAB else "#64748b",
}

SARIF_LEVEL_MAP = {
    "critical": "error",
    "high": "error",
    "medium": "warning",
    "low": "note",
    "info": "note",
}


def _get_severity_str(f) -> str:
    if hasattr(f.severity, "value"):
        return str(f.severity.value).lower()
    return str(f.severity).lower()


def _get_raw_data(f) -> Dict[str, Any]:
    if hasattr(f, 'raw_data') and isinstance(f.raw_data, dict):
        return f.raw_data
    return {}


def generate_html_report(scan) -> str:
    findings = getattr(scan, "findings", [])
    
    # Severity counters
    counts = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
    for f in findings:
        s = _get_severity_str(f)
        if s in counts:
            counts[s] += 1
            
    findings_cards = ""
    for idx, f in enumerate(findings, 1):
        sev = _get_severity_str(f)
        color = SEVERITY_COLORS.get(sev, "#64748b")
        safe_title = html.escape(str(f.title))
        safe_desc = html.escape(str(f.description or ""))
        safe_rem = html.escape(str(f.remediation or "No remediation provided."))
        safe_cat = html.escape(str(f.category or "general"))
        
        rd = _get_raw_data(f)
        cve_val = html.escape(str(rd.get('cve', '') or ''))
        cvss_val = rd.get('cvss_score', '')
        cvss_vec = html.escape(str(rd.get('cvss_vector', '') or ''))
        kev_listed = rd.get('kev_listed', False)
        kev_badge = '<span class="badge badge-kev">Listed in CISA KEV</span>' if kev_listed else ''
        cve_badge = f'<span class="badge badge-cve">{cve_val}</span>' if cve_val else ''
        cvss_badge = f'<span class="badge badge-cvss">CVSS {cvss_val}</span>' if cvss_val else ''
        
        soft_type = html.escape(str(rd.get('software_type', '') or ''))
        plugin_slug = html.escape(str(rd.get('plugin_slug', '') or rd.get('slug', '') or ''))
        det_ver = html.escape(str(rd.get('detected_version', '') or ''))
        patched_ver = html.escape(str(rd.get('patched_version', '') or ''))
        
        meta_items = []
        if soft_type:
            meta_items.append(f"<strong>Type:</strong> {soft_type}")
        if plugin_slug:
            meta_items.append(f"<strong>Slug:</strong> {plugin_slug}")
        if det_ver:
            meta_items.append(f"<strong>Detected Version:</strong> {det_ver}")
        if patched_ver:
            meta_items.append(f"<strong>Patched Version:</strong> {patched_ver}")
        if cvss_vec:
            meta_items.append(f"<strong>CVSS Vector:</strong> {cvss_vec}")
            
        meta_html = " | ".join(meta_items)
        meta_block = f'<div class="finding-meta">{meta_html}</div>' if meta_items else ''

        findings_cards += f"""
        <div class="finding-card severity-{sev}" data-severity="{sev}" data-search="{safe_title.lower()} {safe_desc.lower()} {cve_val.lower()} {safe_cat.lower()}">
            <div class="finding-header">
                <div class="finding-title-group">
                    <span class="badge-sev" style="background:{color}">{sev.upper()}</span>
                    <span class="finding-cat">{safe_cat}</span>
                    <h3 class="finding-title">{safe_title}</h3>
                </div>
                <div class="badge-group">
                    {kev_badge} {cve_badge} {cvss_badge}
                </div>
            </div>
            {meta_block}
            <div class="finding-body">
                <p><strong>Description:</strong> {safe_desc}</p>
                <div class="finding-remediation">
                    <strong>Recommended Action / Remediation:</strong>
                    <p>{safe_rem}</p>
                </div>
            </div>
        </div>
        """

    score_display = scan.score if scan.score is not None else 'N/A'
    grade_display = scan.grade or 'N/A'
    wp_display = 'Yes' if scan.is_wordpress else 'No'
    wp_ver_display = html.escape(str(scan.wp_version or 'Unknown'))
    created_at_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')

    out_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>SentinelWP Enterprise Security Report - {html.escape(str(scan.target_url))}</title>
<style>
:root {{
    --bg: #0f172a;
    --card-bg: #1e293b;
    --text-main: #f8fafc;
    --text-muted: #94a3b8;
    --border: #334155;
    --accent: #38bdf8;
    --critical: #dc2626;
    --high: #ea580c;
    --medium: #d97706;
    --low: #2563eb;
    --info: #64748b;
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; }}
body {{ background: var(--bg); color: var(--text-main); padding: 40px 20px; line-height: 1.6; }}
.container {{ max-width: 1200px; margin: 0 auto; }}
.header {{ border-bottom: 2px solid var(--border); padding-bottom: 20px; margin-bottom: 30px; display: flex; justify-content: space-between; align-items: flex-end; flex-wrap: wrap; gap: 16px; }}
.header h1 {{ font-size: 28px; font-weight: 700; color: #fff; letter-spacing: -0.5px; }}
.header .subtitle {{ color: var(--text-muted); font-size: 14px; margin-top: 4px; }}
.header .report-badge {{ background: rgba(56, 189, 248, 0.1); color: var(--accent); border: 1px solid var(--accent); padding: 6px 12px; border-radius: 6px; font-weight: 600; font-size: 13px; }}

.exec-summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 16px; margin-bottom: 32px; }}
.card {{ background: var(--card-bg); border: 1px solid var(--border); border-radius: 10px; padding: 20px; text-align: center; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); }}
.card-val {{ font-size: 32px; font-weight: 800; color: #fff; margin-bottom: 4px; }}
.card-lbl {{ font-size: 12px; font-weight: 600; text-transform: uppercase; color: var(--text-muted); letter-spacing: 0.5px; }}

.matrix-card {{ background: var(--card-bg); border: 1px solid var(--border); border-radius: 10px; padding: 20px; margin-bottom: 32px; }}
.matrix-title {{ font-size: 16px; font-weight: 600; margin-bottom: 16px; color: #fff; display: flex; align-items: center; justify-content: space-between; }}
.matrix-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; text-align: center; }}
.matrix-item {{ padding: 12px; border-radius: 8px; border: 1px solid var(--border); background: rgba(15, 23, 42, 0.5); }}
.matrix-item .count {{ font-size: 22px; font-weight: 700; }}
.matrix-item .label {{ font-size: 11px; text-transform: uppercase; margin-top: 4px; font-weight: 600; }}

.controls {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 24px; flex-wrap: wrap; gap: 16px; }}
.filter-tabs {{ display: flex; gap: 8px; flex-wrap: wrap; }}
.tab-btn {{ background: var(--card-bg); border: 1px solid var(--border); color: var(--text-muted); padding: 8px 16px; border-radius: 6px; font-size: 13px; font-weight: 600; cursor: pointer; transition: all 0.2s; }}
.tab-btn:hover, .tab-btn.active {{ background: var(--accent); color: #0f172a; border-color: var(--accent); }}
.search-input {{ background: var(--card-bg); border: 1px solid var(--border); color: #fff; padding: 8px 16px; border-radius: 6px; font-size: 13px; width: 280px; }}

.finding-card {{ background: var(--card-bg); border: 1px solid var(--border); border-radius: 10px; padding: 20px; margin-bottom: 16px; transition: border-color 0.2s; }}
.finding-card:hover {{ border-color: #475569; }}
.finding-header {{ display: flex; justify-content: space-between; align-items: flex-start; gap: 16px; margin-bottom: 12px; flex-wrap: wrap; }}
.finding-title-group {{ display: flex; align-items: center; gap: 10px; flex-wrap: wrap; }}
.badge-sev {{ color: #fff; font-size: 11px; font-weight: 700; padding: 4px 8px; border-radius: 4px; letter-spacing: 0.5px; }}
.finding-cat {{ background: #334155; color: #cbd5e1; font-size: 11px; padding: 4px 8px; border-radius: 4px; text-transform: uppercase; font-weight: 600; }}
.finding-title {{ font-size: 16px; font-weight: 600; color: #fff; }}
.badge-group {{ display: flex; gap: 6px; flex-wrap: wrap; }}
.badge {{ font-size: 11px; font-weight: 600; padding: 3px 8px; border-radius: 4px; }}
.badge-kev {{ background: rgba(220, 38, 38, 0.2); color: #f87171; border: 1px solid #dc2626; }}
.badge-cve {{ background: rgba(56, 189, 248, 0.15); color: #38bdf8; border: 1px solid #0284c7; }}
.badge-cvss {{ background: rgba(217, 119, 6, 0.2); color: #fbbf24; border: 1px solid #d97706; }}
.finding-meta {{ font-size: 12px; color: var(--text-muted); background: rgba(15, 23, 42, 0.6); padding: 8px 12px; border-radius: 6px; margin-bottom: 12px; border: 1px solid rgba(255,255,255,0.05); }}
.finding-body {{ font-size: 13px; color: #cbd5e1; }}
.finding-remediation {{ margin-top: 12px; background: rgba(15, 23, 42, 0.8); border-left: 3px solid var(--accent); padding: 10px 14px; border-radius: 0 6px 6px 0; font-size: 13px; }}
.finding-remediation strong {{ color: var(--accent); display: block; margin-bottom: 2px; font-size: 12px; }}

.footer {{ text-align: center; color: var(--text-muted); font-size: 12px; margin-top: 40px; padding-top: 20px; border-top: 1px solid var(--border); }}

@media print {{
    body {{ background: #fff; color: #000; padding: 0; }}
    .finding-card, .card, .matrix-card {{ background: #fff; color: #000; border: 1px solid #ccc; page-break-inside: avoid; }}
    .controls {{ display: none; }}
    .finding-title {{ color: #000; }}
}}
</style>
</head>
<body>
<div class="container">
    <div class="header">
        <div>
            <h1>SentinelWP Security Report</h1>
            <div class="subtitle">Target: <strong>{html.escape(str(scan.target_url))}</strong> | Scan ID: <code>{str(scan.id)[:8]}</code> | Generated: {created_at_str}</div>
        </div>
        <div class="report-badge">Enterprise Edition v3.4</div>
    </div>

    <div class="exec-summary">
        <div class="card">
            <div class="card-val" style="color: {'#ef4444' if scan.score and scan.score > 70 else ('#f59e0b' if scan.score and scan.score > 40 else '#10b981')}">{score_display}</div>
            <div class="card-lbl">Risk Score (0-100)</div>
        </div>
        <div class="card">
            <div class="card-val">{grade_display}</div>
            <div class="card-lbl">Security Grade</div>
        </div>
        <div class="card">
            <div class="card-val">{len(findings)}</div>
            <div class="card-lbl">Total Findings</div>
        </div>
        <div class="card">
            <div class="card-val">{wp_display}</div>
            <div class="card-lbl">WordPress CMS</div>
        </div>
        <div class="card">
            <div class="card-val" style="font-size: 20px; padding-top: 8px;">{wp_ver_display}</div>
            <div class="card-lbl">Core Version</div>
        </div>
    </div>

    <div class="matrix-card">
        <div class="matrix-title">
            <span>Severity Distribution Matrix</span>
            <span style="font-size: 13px; font-weight: normal; color: var(--text-muted);">Active Scan Findings Breakdown</span>
        </div>
        <div class="matrix-grid">
            <div class="matrix-item">
                <div class="count" style="color: var(--critical)">{counts['critical']}</div>
                <div class="label" style="color: var(--critical)">Critical</div>
            </div>
            <div class="matrix-item">
                <div class="count" style="color: var(--high)">{counts['high']}</div>
                <div class="label" style="color: var(--high)">High</div>
            </div>
            <div class="matrix-item">
                <div class="count" style="color: var(--medium)">{counts['medium']}</div>
                <div class="label" style="color: var(--medium)">Medium</div>
            </div>
            <div class="matrix-item">
                <div class="count" style="color: var(--low)">{counts['low']}</div>
                <div class="label" style="color: var(--low)">Low</div>
            </div>
            <div class="matrix-item">
                <div class="count" style="color: var(--info)">{counts['info']}</div>
                <div class="label" style="color: var(--info)">Info</div>
            </div>
        </div>
    </div>

    <div class="controls">
        <div class="filter-tabs">
            <button class="tab-btn active" onclick="filterSev('all', this)">All ({len(findings)})</button>
            <button class="tab-btn" onclick="filterSev('critical', this)">Critical ({counts['critical']})</button>
            <button class="tab-btn" onclick="filterSev('high', this)">High ({counts['high']})</button>
            <button class="tab-btn" onclick="filterSev('medium', this)">Medium ({counts['medium']})</button>
            <button class="tab-btn" onclick="filterSev('low', this)">Low ({counts['low']})</button>
            <button class="tab-btn" onclick="filterSev('info', this)">Info ({counts['info']})</button>
        </div>
        <input type="text" class="search-input" id="searchInput" placeholder="Search findings..." onkeyup="searchFindings()"/>
    </div>

    <div id="findingsContainer">
        {findings_cards if findings_cards else '<div style="text-align:center;padding:40px;color:var(--text-muted)">No security findings detected.</div>'}
    </div>

    <div class="footer">
        Generated by <strong>SentinelWP v3.4.0</strong> — Enterprise WordPress Vulnerability Intelligence &amp; Hardening Platform
    </div>
</div>

<script>
function filterSev(sev, btn) {{
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    
    document.querySelectorAll('.finding-card').forEach(card => {{
        if (sev === 'all' || card.dataset.severity === sev) {{
            card.style.display = 'block';
        }} else {{
            card.style.display = 'none';
        }}
    }});
}}

function searchFindings() {{
    const query = document.getElementById('searchInput').value.toLowerCase();
    document.querySelectorAll('.finding-card').forEach(card => {{
        const text = card.dataset.search || '';
        if (text.includes(query)) {{
            card.style.display = 'block';
        }} else {{
            card.style.display = 'none';
        }}
    }});
}}
</script>
</body>
</html>"""
    return out_html


def generate_json_report(scan) -> str:
    def f_to_dict(f):
        rd = _get_raw_data(f)
        return {
            "id": f.id,
            "severity": _get_severity_str(f),
            "category": f.category,
            "title": f.title,
            "description": f.description,
            "confidence": f.confidence,
            "remediation": f.remediation,
            "reference": f.reference,
            "cve": rd.get("cve"),
            "cvss_score": rd.get("cvss_score"),
            "cvss_vector": rd.get("cvss_vector"),
            "kev_listed": rd.get("kev_listed", False),
            "source": rd.get("source"),
            "affected_range": rd.get("affected_range"),
            "detected_version": rd.get("detected_version"),
            "patched_version": rd.get("patched_version"),
            "plugin_slug": rd.get("plugin_slug"),
            "active_installs": rd.get("active_installs"),
            "match_status": rd.get("match_status"),
        }
    obj = {
        "scan_id": str(scan.id),
        "target_url": str(scan.target_url),
        "score": scan.score,
        "grade": scan.grade,
        "is_wordpress": scan.is_wordpress,
        "wp_version": scan.wp_version,
        "started_at": str(scan.started_at) if scan.started_at is not None else None,
        "completed_at": str(scan.completed_at) if scan.completed_at is not None else None,
        "findings_count": len(scan.findings),
        "findings": [f_to_dict(f) for f in scan.findings],
    }
    return json.dumps(obj, indent=2, ensure_ascii=False)


def generate_sarif_report(scan) -> str:
    rules = []
    results = []
    seen_rule_ids = set()

    for idx, f in enumerate(getattr(scan, "findings", [])):
        sev = _get_severity_str(f)
        level = SARIF_LEVEL_MAP.get(sev, "note")
        rule_id = f"SENTINEL-{f.category.upper()}-{abs(hash(f.title)) % 10000:04d}"

        if rule_id not in seen_rule_ids:
            seen_rule_ids.add(rule_id)
            rules.append({
                "id": rule_id,
                "name": f.title.replace(" ", ""),
                "shortDescription": {"text": f.title},
                "fullDescription": {"text": f.description or ""},
                "help": {"text": f.remediation or "Refer to security hardening guidelines."},
                "properties": {"tags": ["security", "wordpress", f.category]}
            })

        results.append({
            "ruleId": rule_id,
            "level": level,
            "message": {"text": f"{f.title}: {f.description}"},
            "locations": [{
                "physicalLocation": {
                    "artifactLocation": {"uri": scan.target_url},
                    "region": {"startLine": 1}
                }
            }]
        })

    sarif_log = {
        "$schema": "https://raw.githubusercontent.com/oasis-tcs/sarif-spec/master/Schemas/sarif-schema-2.1.0.json",
        "version": "2.1.0",
        "runs": [{
            "tool": {
                "driver": {
                    "name": "SentinelWP",
                    "version": "3.4.0",
                    "informationUri": "https://github.com/melillopietro/SentinelWP",
                    "rules": rules
                }
            },
            "results": results
        }]
    }

    return json.dumps(sarif_log, indent=2, ensure_ascii=False)


def _sanitize_cell(value):
    """Prevent spreadsheet formula injection from external data."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


def generate_excel_report(scan) -> bytes:
    wb = Workbook()
    
    # Sheet 1: Summary Dashboard
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
    ws_sum["A1"] = "SentinelWP — Executive Security Report"
    ws_sum["A1"].font = title_font
    
    ws_sum.append([])
    ws_sum.append(["Target URL", str(scan.target_url)])
    ws_sum.append(["Scan ID", str(scan.id)])
    ws_sum.append(["Report Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    ws_sum.append(["Risk Score", scan.score if scan.score is not None else "N/A"])
    ws_sum.append(["Security Grade", scan.grade or "N/A"])
    ws_sum.append(["WordPress CMS Detected", "Yes" if scan.is_wordpress else "No"])
    ws_sum.append(["WordPress Core Version", scan.wp_version or "Unknown"])
    ws_sum.append(["Total Findings", len(scan.findings)])
    
    # Format Summary block
    lbl_font = Font(name="Segoe UI", size=11, bold=True, color="334155")
    val_font = Font(name="Segoe UI", size=11, color="0F172A")
    for row in ws_sum.iter_rows(min_row=3, max_row=10, min_col=1, max_col=2):
        row[0].font = lbl_font
        row[1].font = val_font
        
    ws_sum.append([])
    ws_sum.append(["Severity Breakdown", "Count"])
    ws_sum.cell(row=12, column=1).font = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    ws_sum.cell(row=12, column=2).font = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    
    counts = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
    for f in scan.findings:
        s = _get_severity_str(f)
        if s in counts:
            counts[s] += 1
            
    for sev, count in counts.items():
        ws_sum.append([sev.upper(), count])

    # Sheet 2: Findings Detail
    ws_detail = wb.create_sheet(title="Findings Detail")
    headers = [
        "Severity", "Category", "Title", "CVE", "CVSS Score", "CVSS Vector",
        "CISA KEV Listed", "Software Type", "Slug", "Detected Version",
        "Patched Version", "Description", "Remediation Action"
    ]
    ws_detail.append(headers)
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    for cell in ws_detail[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for f in scan.findings:
        sev = _get_severity_str(f)
        rd = _get_raw_data(f)
        row_data = [
            sev.upper(),
            f.category,
            f.title,
            rd.get("cve", ""),
            rd.get("cvss_score", ""),
            rd.get("cvss_vector", ""),
            "YES" if rd.get("kev_listed") else "NO",
            rd.get("software_type", ""),
            rd.get("plugin_slug", "") or rd.get("slug", ""),
            rd.get("detected_version", ""),
            rd.get("patched_version", ""),
            f.description,
            f.remediation
        ]
        ws_detail.append([_sanitize_cell(c) for c in row_data])

    # Auto-fit column widths across sheets
    for ws in [ws_sum, ws_detail]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(50, max(12, max_len + 3))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if HAS_REPORTLAB:
    class NumberedCanvas(canvas.Canvas):
        """
        Two-pass canvas to dynamically compute and render total page count ('Page X of Y')
        along with enterprise running headers and footers.
        """
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_page_decorations(num_pages)
                super().showPage()
            super().save()

        def draw_page_decorations(self, page_count):
            self.saveState()
            self.setFont("Helvetica", 8)
            self.setFillColor(colors.HexColor("#64748b"))

            # Running Header on page 2+
            if self._pageNumber > 1:
                self.drawString(2 * cm, 28.2 * cm, "SentinelWP — Executive Security Report")
                self.setStrokeColor(colors.HexColor("#cbd5e1"))
                self.setLineWidth(0.5)
                self.line(2 * cm, 28.0 * cm, 19.0 * cm, 28.0 * cm)

            # Running Footer on all pages
            self.setStrokeColor(colors.HexColor("#cbd5e1"))
            self.setLineWidth(0.5)
            self.line(2 * cm, 1.8 * cm, 19.0 * cm, 1.8 * cm)

            footer_text = "SentinelWP v3.4.0 — Confidential Security Assessment Report"
            page_text = f"Page {self._pageNumber} of {page_count}"
            
            self.drawString(2 * cm, 1.3 * cm, footer_text)
            self.drawRightString(19.0 * cm, 1.3 * cm, page_text)
            self.restoreState()


def generate_pdf_report(scan) -> bytes:
    if not HAS_REPORTLAB:
        raise RuntimeError("ReportLab is required for PDF generation.")

    buf = io.BytesIO()
    # A4 Page size: 21.0cm x 29.7cm. Margins: 2.0cm left/right -> Printable Width = 17.0cm
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2.0 * cm,
        rightMargin=2.0 * cm,
        topMargin=2.0 * cm,
        bottomMargin=2.2 * cm
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Enterprise Typography
    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0f172a"),
        alignment=TA_LEFT,
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        "DocSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=12
    )

    h2_style = ParagraphStyle(
        "SectionH2",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=17,
        textColor=colors.HexColor("#0f172a"),
        spaceBefore=14,
        spaceAfter=8
    )

    body_style = ParagraphStyle(
        "BodyDark",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#334155")
    )

    cell_hdr_style = ParagraphStyle(
        "CellHdr",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=TA_LEFT
    )

    cell_body_style = ParagraphStyle(
        "CellBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#0f172a")
    )

    cell_body_bold = ParagraphStyle(
        "CellBodyBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#0f172a")
    )

    elems = []

    # Title & Subtitle Header
    elems.append(Paragraph("SentinelWP — WordPress Security Assessment", title_style))
    created_at_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    elems.append(Paragraph(
        f"<b>Target URL:</b> {html.escape(str(scan.target_url))} &nbsp;|&nbsp; "
        f"<b>Scan ID:</b> <code>{str(scan.id)[:8]}</code> &nbsp;|&nbsp; "
        f"<b>Generated:</b> {created_at_str}",
        subtitle_style
    ))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#0f172a"), spaceAfter=12))

    # Executive Summary Card Table
    elems.append(Paragraph("Executive Summary", h2_style))
    score_str = str(scan.score) if scan.score is not None else "N/A"
    grade_str = str(scan.grade or "N/A")
    wp_str = "Yes" if scan.is_wordpress else "No"
    wp_ver = str(scan.wp_version or "Unknown")

    summary_rows = [
        [
            Paragraph("<b>Overall Risk Score</b>", cell_body_style),
            Paragraph(f"<b>{score_str} / 100</b>", cell_body_bold),
            Paragraph("<b>Security Grade</b>", cell_body_style),
            Paragraph(f"<b>{grade_str}</b>", cell_body_bold),
        ],
        [
            Paragraph("<b>WordPress CMS</b>", cell_body_style),
            Paragraph(wp_str, cell_body_style),
            Paragraph("<b>Core WP Version</b>", cell_body_style),
            Paragraph(wp_ver, cell_body_style),
        ],
        [
            Paragraph("<b>Total Findings</b>", cell_body_style),
            Paragraph(str(len(scan.findings)), cell_body_bold),
            Paragraph("<b>Scanner Engine</b>", cell_body_style),
            Paragraph("SentinelWP v3.4 Enterprise", cell_body_style),
        ]
    ]

    sum_table = Table(summary_rows, colWidths=[4.0*cm, 4.5*cm, 4.0*cm, 4.5*cm])
    sum_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems.append(sum_table)
    elems.append(Spacer(1, 10))

    # Severity Matrix Breakdown
    counts = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
    for f in scan.findings:
        s = _get_severity_str(f)
        if s in counts:
            counts[s] += 1

    matrix_data = [
        [
            Paragraph("<b>CRITICAL</b>", ParagraphStyle("C", parent=cell_hdr_style, alignment=TA_CENTER)),
            Paragraph("<b>HIGH</b>", ParagraphStyle("H", parent=cell_hdr_style, alignment=TA_CENTER)),
            Paragraph("<b>MEDIUM</b>", ParagraphStyle("M", parent=cell_hdr_style, alignment=TA_CENTER)),
            Paragraph("<b>LOW</b>", ParagraphStyle("L", parent=cell_hdr_style, alignment=TA_CENTER)),
            Paragraph("<b>INFO</b>", ParagraphStyle("I", parent=cell_hdr_style, alignment=TA_CENTER)),
        ],
        [
            Paragraph(f"<b>{counts['critical']}</b>", ParagraphStyle("C1", parent=cell_body_style, alignment=TA_CENTER, fontSize=11)),
            Paragraph(f"<b>{counts['high']}</b>", ParagraphStyle("H1", parent=cell_body_style, alignment=TA_CENTER, fontSize=11)),
            Paragraph(f"<b>{counts['medium']}</b>", ParagraphStyle("M1", parent=cell_body_style, alignment=TA_CENTER, fontSize=11)),
            Paragraph(f"<b>{counts['low']}</b>", ParagraphStyle("L1", parent=cell_body_style, alignment=TA_CENTER, fontSize=11)),
            Paragraph(f"<b>{counts['info']}</b>", ParagraphStyle("I1", parent=cell_body_style, alignment=TA_CENTER, fontSize=11)),
        ]
    ]

    mat_table = Table(matrix_data, colWidths=[3.4*cm, 3.4*cm, 3.4*cm, 3.4*cm, 3.4*cm])
    mat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#dc2626")),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#ea580c")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#d97706")),
        ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#2563eb")),
        ("BACKGROUND", (4, 0), (4, 0), colors.HexColor("#64748b")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#ffffff")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems.append(mat_table)
    elems.append(Spacer(1, 14))

    # Findings Detail Table Section
    elems.append(Paragraph("Detailed Vulnerability & Finding Log", h2_style))

    if not scan.findings:
        elems.append(Paragraph("No vulnerability findings or exposures were detected during this scan.", body_style))
    else:
        # Table columns total 17.0cm:
        # [Severity: 2.2cm, Category: 2.5cm, Title & CVE: 4.8cm, Description & Remediation: 7.5cm]
        findings_table_data = [
            [
                Paragraph("Severity", cell_hdr_style),
                Paragraph("Category", cell_hdr_style),
                Paragraph("Title & CVE", cell_hdr_style),
                Paragraph("Description & Action", cell_hdr_style)
            ]
        ]

        for f in scan.findings:
            sev = _get_severity_str(f)
            rd = _get_raw_data(f)
            
            cve_str = rd.get("cve", "")
            cvss_str = str(rd.get("cvss_score", "")) if rd.get("cvss_score") else ""
            kev_str = " (CISA KEV Listed)" if rd.get("kev_listed") else ""
            
            # Severity Flowable
            sev_p = Paragraph(f"<font color='{SEVERITY_COLORS.get(sev, '#000')}'><b>{sev.upper()}</b></font>", cell_body_bold)
            
            # Category Flowable (Clean label for vuln intel)
            cat_name = f.category or "general"
            if cat_name == "vulnerability_intelligence":
                cat_name = "threat_intel"
            cat_p = Paragraph(html.escape(cat_name), cell_body_style)
            
            # Title & CVE Flowable
            cve_line = f"<br/><font color='#2563eb'><b>CVE:</b> {html.escape(cve_str)}</font>" if cve_str else ""
            cvss_line = f" | <b>CVSS:</b> {cvss_str}" if cvss_str else ""
            kev_line = f"<br/><font color='#dc2626'><b>{kev_str}</b></font>" if kev_str else ""
            title_text = f"<b>{html.escape(str(f.title))}</b>{cve_line}{cvss_line}{kev_line}"
            title_p = Paragraph(title_text, cell_body_style)
            
            # Description & Remediation Flowable (Strip raw python range dict dumps)
            raw_desc = str(f.description or "")
            clean_desc = re.sub(r"\s*\(range:\s*\[\{.*?\}\]\)", "", raw_desc, flags=re.DOTALL)
            desc_text = html.escape(clean_desc)
            rem_text = html.escape(str(f.remediation or "Follow security hardening practices."))
            
            det_ver = rd.get("detected_version")
            patched_ver = rd.get("patched_version")
            ver_meta = ""
            if det_ver or patched_ver:
                ver_meta = f"<br/><font color='#475569'><b>Detected:</b> {det_ver or 'N/A'} | <b>Patched:</b> {patched_ver or 'N/A'}</font>"
                
            body_text = f"{desc_text}{ver_meta}<br/><font color='#0284c7'><b>Action:</b> {rem_text}</font>"
            body_p = Paragraph(body_text, cell_body_style)
            
            findings_table_data.append([sev_p, cat_p, title_p, body_p])

        ft = Table(findings_table_data, colWidths=[1.9*cm, 2.6*cm, 4.8*cm, 7.7*cm], repeatRows=1)
        ft.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        elems.append(ft)

    elems.append(Spacer(1, 16))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cbd5e1")))
    elems.append(Paragraph(
        "<b>SentinelWP Security Sentinel</b> — Continuous WordPress Threat Intelligence &amp; Automated Vulnerability Management",
        subtitle_style
    ))

    doc.build(elems, canvasmaker=NumberedCanvas)
    return buf.getvalue()
